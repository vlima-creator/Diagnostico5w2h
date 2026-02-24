# -*- coding: utf-8 -*-
"""
Aplicação Streamlit - Diagnóstico 5W2H para Reuniões de Start
Versão 2.0 - Banco de Dados Otimizado
Permite capturar ações durante reunião e gerar plano 5W2H automático
"""

import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
from datetime import datetime, timedelta
import json
import io
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Image
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================================
# CONFIGURAÇÃO DA PÁGINA
# ============================================================================

st.set_page_config(
    page_title="Diagnóstico 5W2H",
    page_icon="📋",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ============================================================================
# BANCO DE DADOS DE AÇÕES OTIMIZADO (v2.0)
# ============================================================================

BANCO_ACOES = [
    {
        "id": 1,
        "acao": "Ajustar precificacao",
        "categoria": "Precificacao",
        "what": "Revisar e ajustar preços dos SKUs foco com base em margem, concorrência e regras de frete.",
        "why": "Melhorar conversão sem perder margem, reduzir perda para concorrentes e evitar preço abaixo do mínimo.",
        "where": "No(s) canal(is) priorizado(s) no ciclo (ex.: Mercado Livre, Shopee, Amazon, site).",
        "how": "1) Selecionar SKUs foco (top vendas e top visitas)\n2) Levantar custo total e preço mínimo\n3) Comparar com concorrentes e elasticidade\n4) Definir regras por faixa e por canal\n5) Atualizar preços e monitorar por 7 dias\n6) Ajustar a partir de conversão e margem",
        "indicadores": "Margem, conversão, visitas, buy box, GMV, ticket médio",
        "dia_inicio_padrao": 1,
        "duracao_dias": 7,
        "custo_padrao": 0,
        "impacto_padrao": 4,
        "esforco_padrao": 3
    },
    {
        "id": 2,
        "acao": "Contratar colaborador",
        "categoria": "Pessoas",
        "what": "Abrir vaga, selecionar e contratar colaborador para função definida.",
        "why": "Criar capacidade de execução, tirar gargalos e sustentar crescimento.",
        "where": "Time interno, remoto ou híbrido, conforme necessidade.",
        "how": "1) Definir escopo e metas da função\n2) Escrever descrição da vaga\n3) Divulgar e captar candidatos\n4) Triagem e entrevistas\n5) Teste prático\n6) Proposta e admissão\n7) Onboarding e metas de 30 dias",
        "indicadores": "Tempo de preenchimento, ramp-up, entregas no 30º dia, qualidade de execução",
        "dia_inicio_padrao": 3,
        "duracao_dias": 21,
        "custo_padrao": 0,
        "impacto_padrao": 4,
        "esforco_padrao": 4
    },
    {
        "id": 3,
        "acao": "Implementar ERP (sistema de integracao)",
        "categoria": "Sistemas",
        "what": "Selecionar, configurar e implementar um ERP para integracao de dados de vendas, estoque e financeiro.",
        "why": "Centralizar informacoes, reduzir erros manuais, melhorar visibilidade operacional e facilitar decisoes baseadas em dados.",
        "where": "Sistemas internos da empresa, integrando marketplaces, estoque e financeiro.",
        "how": "1) Definir requisitos e necessidades do negocio\n2) Pesquisar e avaliar opcoes de ERP (Bling, Omie, SAP, etc)\n3) Negociar contrato e implementacao\n4) Configurar modulos (vendas, estoque, financeiro)\n5) Integrar com marketplaces e canais\n6) Treinar equipe no novo sistema\n7) Monitorar e ajustar conforme necessario",
        "indicadores": "Tempo de implementacao, taxa de adocao, reducao de erros, tempo de resposta, acuracia de estoque",
        "dia_inicio_padrao": 1,
        "duracao_dias": 30,
        "custo_padrao": 0,
        "impacto_padrao": 5,
        "esforco_padrao": 5
    },
    {
        "id": 4,
        "acao": "Utilizacao de canal (otimizacao do canal atual)",
        "categoria": "Canais",
        "what": "Revisar setup do canal atual e ajustar catálogo, preço, reputação, prazos e operação.",
        "why": "Aumentar conversão e relevância usando melhor o canal que já existe.",
        "where": "Canal atual prioritário (ex.: Mercado Livre).",
        "how": "1) Diagnóstico: reputação, métricas, prazos, políticas\n2) Ajustar cadastro e conteúdo de produtos\n3) Revisar preços e frete\n4) Ajustar estoque e ruptura\n5) Implementar melhorias identificadas\n6) Monitorar impacto por 7 dias\n7) Ajustar conforme resultados",
        "indicadores": "Conversão, reputação, ruptura, visitas, tempo de envio, devoluções",
        "dia_inicio_padrao": 1,
        "duracao_dias": 14,
        "custo_padrao": 0,
        "impacto_padrao": 4,
        "esforco_padrao": 3
    },
    {
        "id": 5,
        "acao": "Entrada em um novo canal",
        "categoria": "Canais",
        "what": "Abrir e operar um novo canal de vendas, com base em catálogo e capacidade logística.",
        "why": "Diversificar receita e capturar demanda onde o público já compra.",
        "where": "Novo canal escolhido (ex.: Amazon, Magalu, Shein, B2W, TikTok Shop).",
        "how": "1) Validar requisitos e custos do canal\n2) Escolher sortimento inicial (top SKUs)\n3) Preparar integração e cadastro\n4) Subir anúncios e políticas\n5) Treinar equipe no novo canal\n6) Primeiras vendas e ajustes\n7) Escalar com campanhas e reposição",
        "indicadores": "GMV por canal, CAC, conversão, margem, SLA, cancelamentos",
        "dia_inicio_padrao": 12,
        "duracao_dias": 21,
        "custo_padrao": 0,
        "impacto_padrao": 4,
        "esforco_padrao": 4
    },
    {
        "id": 6,
        "acao": "Ativar fulfillment",
        "categoria": "Operacao",
        "what": "Ativar modelo de fulfillment (estoque em CD) para SKUs elegíveis.",
        "why": "Ganhar prazo e relevância, reduzir falhas de expedição e melhorar conversão.",
        "where": "Programa do canal (ex.: Full, FBA, etc.) e centros de distribuição.",
        "how": "1) Selecionar SKUs (giro x margem x dimensões)\n2) Conferir custos e regras do programa\n3) Preparar etiquetagem e envio\n4) Enviar lote piloto\n5) Acompanhar nível de serviço e vendas\n6) Expandir sortimento conforme resultados",
        "indicadores": "Conversão, tempo de entrega, cancelamentos, devoluções, GMV, custo logística",
        "dia_inicio_padrao": 8,
        "duracao_dias": 23,
        "custo_padrao": 0,
        "impacto_padrao": 5,
        "esforco_padrao": 4
    },
    {
        "id": 7,
        "acao": "Ativar publicidade",
        "categoria": "Midia",
        "what": "Ativar campanhas pagas no canal (busca, produto, vitrine) com estrutura básica.",
        "why": "Gerar demanda previsível, acelerar vendas e aprender quais produtos respondem melhor.",
        "where": "Painel de anúncios do canal e/ou ferramentas integradas.",
        "how": "1) Definir objetivo e verba inicial\n2) Separar campanhas por objetivo (tráfego, conversão, marca)\n3) Escolher SKUs e palavras-chave\n4) Subir campanhas e anúncios\n5) Monitorar diário por 7 dias\n6) Ajustar lances, negativos e criativos",
        "indicadores": "ROAS/ACOS, CPC, conversão, share de impressão, GMV incremental",
        "dia_inicio_padrao": 5,
        "duracao_dias": 26,
        "custo_padrao": 0,
        "impacto_padrao": 5,
        "esforco_padrao": 4
    },
    {
        "id": 8,
        "acao": "Trabalhar com promocoes",
        "categoria": "Comercial",
        "what": "Planejar e executar promoções (cupons, descontos, kit) em SKUs estratégicos.",
        "why": "Aumentar volume e visibilidade em períodos de maior competição.",
        "where": "Calendário promocional do canal e página de ofertas.",
        "how": "1) Definir SKUs e limites de margem\n2) Escolher mecânica (cupom, desconto, kit)\n3) Criar calendário e comunicar internamente\n4) Rodar promo e monitorar\n5) Ajustar estoque e preços\n6) Avaliar pós-mortem e documentar resultados",
        "indicadores": "GMV, margem, conversão, ruptura, novos clientes, ranking",
        "dia_inicio_padrao": 10,
        "duracao_dias": 14,
        "custo_padrao": 0,
        "impacto_padrao": 4,
        "esforco_padrao": 2
    },
    {
        "id": 9,
        "acao": "Melhorar conteudo e cadastro",
        "categoria": "Catalogo",
        "what": "Padronizar títulos, imagens, atributos e descrições para aumentar relevância e conversão.",
        "why": "Reduzir atrito de compra e aumentar qualidade de anúncio.",
        "where": "Catálogo do canal e integrador (se houver).",
        "how": "1) Definir padrão por categoria\n2) Corrigir top 20 SKUs\n3) Replicar padrão no restante\n4) Auditar atributos obrigatórios\n5) Testar imagens e títulos\n6) Revisar mensalmente",
        "indicadores": "Conversão, visitas, reclamações, taxa de perguntas, índice de qualidade",
        "dia_inicio_padrao": 2,
        "duracao_dias": 14,
        "custo_padrao": 0,
        "impacto_padrao": 4,
        "esforco_padrao": 3
    },
    {
        "id": 10,
        "acao": "Implantar rotina de indicadores",
        "categoria": "Gestao",
        "what": "Criar rotina semanal de acompanhamento e decisão com base em indicadores.",
        "why": "Aumentar velocidade de decisão e manter foco no que traz resultado.",
        "where": "Reunião semanal e painel (Sheets/BI).",
        "how": "1) Definir KPIs e metas\n2) Montar painel simples (Sheets/BI)\n3) Ritual semanal: revisar, decidir, delegar\n4) Registrar plano de ação\n5) Documentar decisões e ações tomadas\n6) Fazer follow-up das ações na semana seguinte\n7) Revisar resultados em 30 dias",
        "indicadores": "GMV, margem, conversão, ruptura, ROAS, SLA, devoluções",
        "dia_inicio_padrao": 1,
        "duracao_dias": 30,
        "custo_padrao": 0,
        "impacto_padrao": 5,
        "esforco_padrao": 2
    },
    {
        "id": 11,
        "acao": "Rotinas e processos",
        "categoria": "Gestao",
        "what": "Estruturar e documentar rotinas operacionais e estratégicas.",
        "why": "Garantir padrão, previsibilidade e escalabilidade da operação.",
        "where": "Operação geral da empresa.",
        "how": "1) Mapear processos atuais\n2) Identificar gargalos e ineficiências\n3) Documentar fluxo ideal\n4) Definir responsáveis\n5) Treinar equipe nos novos processos\n6) Monitorar aderência por 2 semanas\n7) Ajustar conforme feedback",
        "indicadores": "Tempo de execução, retrabalho, erros operacionais",
        "dia_inicio_padrao": 1,
        "duracao_dias": 30,
        "custo_padrao": 0,
        "impacto_padrao": 5,
        "esforco_padrao": 3
    },
    {
        "id": 12,
        "acao": "Analisar a curva ABC",
        "categoria": "Gestao",
        "what": "Classificar produtos por representatividade de faturamento e margem.",
        "why": "Priorizar foco nos produtos que realmente movem o resultado.",
        "where": "Relatórios de vendas e ERP.",
        "how": "1) Exportar vendas (últimos 90 dias)\n2) Classificar por faturamento\n3) Separar A, B e C\n4) Definir estratégia por curva\n5) Criar plano de ação baseado na curva ABC\n6) Comunicar resultados à equipe\n7) Revisar mensalmente",
        "indicadores": "GMV por SKU, margem, giro",
        "dia_inicio_padrao": 1,
        "duracao_dias": 5,
        "custo_padrao": 0,
        "impacto_padrao": 4,
        "esforco_padrao": 2
    },
    {
        "id": 13,
        "acao": "Gestao de Compras (Mix de Produtos)",
        "categoria": "Compras",
        "what": "Definir mix ideal baseado em giro e margem.",
        "why": "Evitar ruptura e excesso de estoque.",
        "where": "ERP e relatórios de estoque.",
        "how": "1) Cruzar curva ABC com estoque\n2) Identificar ruptura e excesso\n3) Planejar reposição\n4) Negociar fornecedores\n5) Implementar novo mix no sistema\n6) Monitorar resultados por 7 dias\n7) Ajustar conforme performance",
        "indicadores": "Ruptura, giro, cobertura de estoque",
        "dia_inicio_padrao": 3,
        "duracao_dias": 15,
        "custo_padrao": 0,
        "impacto_padrao": 5,
        "esforco_padrao": 4
    },
    {
        "id": 14,
        "acao": "Gestao de Atendimento",
        "categoria": "Operacao",
        "what": "Padronizar e monitorar atendimento ao cliente.",
        "why": "Melhorar reputação e conversão.",
        "where": "Canal de atendimento do marketplace.",
        "how": "1) Criar scripts padrão\n2) Definir SLA (tempo de resposta)\n3) Monitorar tempo de resposta\n4) Revisar feedbacks e reclamações\n5) Implementar sistema de monitoramento\n6) Fazer reunião de feedback com equipe\n7) Treinar conforme necessidade",
        "indicadores": "Tempo resposta, reputação, NPS",
        "dia_inicio_padrao": 1,
        "duracao_dias": 30,
        "custo_padrao": 0,
        "impacto_padrao": 4,
        "esforco_padrao": 3
    },
    {
        "id": 15,
        "acao": "Padronizacao de Anuncios",
        "categoria": "Catalogo",
        "what": "Criar padrão de títulos, imagens e descrições.",
        "why": "Aumentar conversão e qualidade dos anúncios.",
        "where": "Anúncios ativos no marketplace.",
        "how": "1) Definir modelo padrão\n2) Ajustar top SKUs\n3) Replicar modelo\n4) Revisar atributos\n5) Testar padrão com A/B testing\n6) Monitorar performance por 7 dias\n7) Documentar aprendizados",
        "indicadores": "Conversão, visitas, índice de qualidade",
        "dia_inicio_padrao": 2,
        "duracao_dias": 14,
        "custo_padrao": 0,
        "impacto_padrao": 4,
        "esforco_padrao": 3
    },
    {
        "id": 16,
        "acao": "Padronizacao de Cadastros",
        "categoria": "Catalogo",
        "what": "Padronizar atributos e informações técnicas dos produtos.",
        "why": "Evitar erros e melhorar indexação.",
        "where": "ERP e marketplace.",
        "how": "1) Revisar atributos obrigatórios\n2) Criar checklist de validação\n3) Corrigir inconsistências\n4) Corrigir inconsistências identificadas\n5) Validar integração com marketplace\n6) Documentar padrão para referência futura\n7) Treinar equipe no novo padrão",
        "indicadores": "Erros de integração, qualidade de cadastro",
        "dia_inicio_padrao": 2,
        "duracao_dias": 14,
        "custo_padrao": 0,
        "impacto_padrao": 4,
        "esforco_padrao": 3
    },
    {
        "id": 17,
        "acao": "Ads e Campanhas",
        "categoria": "Midia",
        "what": "Estruturar campanhas de anúncios pagos.",
        "why": "Gerar demanda previsível e escalar vendas.",
        "where": "Painel de anúncios do canal.",
        "how": "1) Definir verba e objetivos\n2) Criar campanhas por objetivo\n3) Monitorar diário\n4) Ajustar palavras e lances\n5) Definir KPIs e metas de performance\n6) Criar dashboard de monitoramento\n7) Escalar vencedores",
        "indicadores": "ROAS, ACOS, CPC, GMV",
        "dia_inicio_padrao": 5,
        "duracao_dias": 26,
        "custo_padrao": 0,
        "impacto_padrao": 5,
        "esforco_padrao": 4
    },
    {
        "id": 18,
        "acao": "Expedicao (PICK & PACK)",
        "categoria": "Operacao",
        "what": "Organizar processo de separação e envio.",
        "why": "Reduzir erros e atrasos logísticos.",
        "where": "Centro de distribuição interno.",
        "how": "1) Mapear fluxo atual\n2) Criar padrão de separação\n3) Organizar layout do CD\n4) Treinar equipe\n5) Implementar sistema de rastreamento\n6) Fazer auditoria de qualidade\n7) Monitorar SLA",
        "indicadores": "Erros de envio, prazo, cancelamentos",
        "dia_inicio_padrao": 1,
        "duracao_dias": 20,
        "custo_padrao": 0,
        "impacto_padrao": 4,
        "esforco_padrao": 4
    },
    {
        "id": 19,
        "acao": "Entrada Produtos",
        "categoria": "Operacao",
        "what": "Padronizar recebimento e cadastro de novos produtos.",
        "why": "Evitar divergências de estoque.",
        "where": "Estoque e ERP.",
        "how": "1) Conferência física\n2) Cadastro correto no ERP\n3) Validação de custo\n4) Treinar equipe no novo processo\n5) Monitorar aderência\n6) Fazer auditoria de qualidade\n7) Documentar procedimento",
        "indicadores": "Erros de estoque, divergências",
        "dia_inicio_padrao": 3,
        "duracao_dias": 10,
        "custo_padrao": 0,
        "impacto_padrao": 4,
        "esforco_padrao": 3
    },
    {
        "id": 20,
        "acao": "Compras",
        "categoria": "Compras",
        "what": "Planejar e executar compras estratégicas.",
        "why": "Garantir abastecimento sem excesso de capital parado.",
        "where": "Fornecedores e ERP.",
        "how": "1) Analisar giro\n2) Definir necessidade\n3) Negociar com fornecedores\n4) Negociar prazos e condições de pagamento\n5) Confirmar data de entrega\n6) Acompanhar recebimento\n7) Validar qualidade e quantidade",
        "indicadores": "Cobertura estoque, margem, giro",
        "dia_inicio_padrao": 5,
        "duracao_dias": 15,
        "custo_padrao": 0,
        "impacto_padrao": 4,
        "esforco_padrao": 3
    }
]

CONFIG = {
    "peso_impacto": 10,
    "peso_esforco": 2,
    "dias_ciclo": 30
}

# ============================================================================
# FUNÇÕES UTILITÁRIAS
# ============================================================================

def calcular_score(impacto, esforco):
    """Calcula o score da ação: (Impacto × Peso) - (Esforço × Peso)"""
    return (impacto * CONFIG["peso_impacto"]) - (esforco * CONFIG["peso_esforco"])

def obter_acao_por_nome(nome_acao):
    """Busca uma ação no banco de dados pelo nome"""
    for acao in BANCO_ACOES:
        if acao["acao"].lower() == nome_acao.lower():
            return acao
    return None

def criar_dataframe_5w2h(cliente, acao_dict, responsavel, notas, data_inicio):
    """Cria um DataFrame com o plano 5W2H preenchido"""
    data_fim = data_inicio + timedelta(days=acao_dict["duracao_dias"])
    score = calcular_score(acao_dict["impacto_padrao"], acao_dict["esforco_padrao"])
    
    return {
        "Cliente": cliente,
        "Acao": acao_dict["acao"],
        "Categoria": acao_dict["categoria"],
        "What": acao_dict["what"],
        "Why": acao_dict["why"],
        "Where": acao_dict["where"],
        "Data_inicio": data_inicio.strftime("%d/%m/%Y"),
        "Duracao_dias": acao_dict["duracao_dias"],
        "Data_fim": data_fim.strftime("%d/%m/%Y"),
        "Responsavel": responsavel,
        "How": acao_dict["how"],
        "HowMuch": acao_dict["custo_padrao"],
        "Impacto": acao_dict["impacto_padrao"],
        "Esforco": acao_dict["esforco_padrao"],
        "Score": score,
        "Notas": notas,
        "Status": "Planejado"
    }

def gerar_gantt_chart(plano_data):
    """Gera um gráfico Gantt com as ações do plano"""
    df = pd.DataFrame([plano_data])
    
    fig = go.Figure()
    
    fig.add_trace(go.Bar(
        y=[df["Acao"].values[0]],
        x=[df["Duracao_dias"].values[0]],
        orientation='h',
        marker=dict(color='#1E3A8A'),
        name='Duração'
    ))
    
    fig.update_layout(
        title="Timeline da Ação (30 dias)",
        xaxis_title="Dias",
        yaxis_title="Ação",
        height=300,
        showlegend=False,
        hovermode='closest'
    )
    
    return fig

def gerar_relatorio_pdf(cliente, plano_data):
    """Gera um relatório em PDF com o plano 5W2H"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    story = []
    styles = getSampleStyleSheet()
    
    # Estilos customizados
    titulo_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1E3A8A'),
        spaceAfter=30,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#1E3A8A'),
        spaceAfter=12,
        fontName='Helvetica-Bold'
    )
    
    # Título
    story.append(Paragraph("DIAGNÓSTICO 5W2H", titulo_style))
    story.append(Paragraph(f"Cliente: <b>{cliente}</b>", styles['Normal']))
    story.append(Paragraph(f"Data: <b>{datetime.now().strftime('%d/%m/%Y')}</b>", styles['Normal']))
    story.append(Spacer(1, 0.3*inch))
    
    # Seção 5W2H
    story.append(Paragraph("PLANO DE AÇÃO", heading_style))
    
    # Tabela com os 5W2H
    data_table = [
        ["Campo", "Descrição"],
        ["WHAT (O quê)", plano_data.get("What", "")],
        ["WHY (Por quê)", plano_data.get("Why", "")],
        ["WHERE (Onde)", plano_data.get("Where", "")],
        ["WHEN (Quando)", f"De {plano_data.get('Data_inicio')} a {plano_data.get('Data_fim')}"],
        ["WHO (Quem)", plano_data.get("Responsavel", "")],
        ["HOW (Como)", plano_data.get("How", "")],
        ["HOW MUCH (Quanto)", f"R$ {plano_data.get('HowMuch', 0):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")],
    ]
    
    table = Table(data_table, colWidths=[1.5*inch, 4.5*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A8A')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    
    story.append(table)
    story.append(Spacer(1, 0.3*inch))
    
    # Indicadores
    story.append(Paragraph("INDICADORES DE SUCESSO", heading_style))
    story.append(Paragraph(
        f"<b>Impacto:</b> {plano_data.get('Impacto', 0)}/5 | "
        f"<b>Esforço:</b> {plano_data.get('Esforco', 0)}/5 | "
        f"<b>Score:</b> {plano_data.get('Score', 0)}",
        styles['Normal']
    ))
    story.append(Spacer(1, 0.2*inch))
    
    # Notas
    if plano_data.get("Notas"):
        story.append(Paragraph("NOTAS DA REUNIÃO", heading_style))
        story.append(Paragraph(plano_data.get("Notas", ""), styles['Normal']))
    
    # Build PDF
    doc.build(story)
    buffer.seek(0)
    return buffer

def gerar_relatorio_excel(cliente, plano_data):
    """Gera um relatório em Excel com o plano 5W2H"""
    buffer = io.BytesIO()
    
    # Criar workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plano 5W2H"
    
    # Estilos
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    ws['A1'] = "DIAGNÓSTICO 5W2H"
    ws['A1'].font = Font(bold=True, size=14, color="1E3A8A")
    ws.merge_cells('A1:B1')
    
    ws['A2'] = f"Cliente: {cliente}"
    ws['A3'] = f"Data: {datetime.now().strftime('%d/%m/%Y')}"
    
    # Dados do 5W2H
    row = 5
    campos = [
        ("WHAT (O quê)", plano_data.get("What", "")),
        ("WHY (Por quê)", plano_data.get("Why", "")),
        ("WHERE (Onde)", plano_data.get("Where", "")),
        ("WHEN (Quando)", f"De {plano_data.get('Data_inicio')} a {plano_data.get('Data_fim')}"),
        ("WHO (Quem)", plano_data.get("Responsavel", "")),
        ("HOW (Como)", plano_data.get("How", "")),
        ("HOW MUCH (Quanto)", f"R$ {plano_data.get('HowMuch', 0):,.2f}"),
    ]
    
    for campo, valor in campos:
        ws[f'A{row}'] = campo
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws[f'A{row}'].border = border
        
        ws[f'B{row}'] = valor
        ws[f'B{row}'].border = border
        ws[f'B{row}'].alignment = Alignment(wrap_text=True)
        
        row += 1
    
    # Indicadores
    row += 1
    ws[f'A{row}'] = "INDICADORES"
    ws[f'A{row}'].font = Font(bold=True, size=12, color="1E3A8A")
    
    row += 1
    ws[f'A{row}'] = "Impacto"
    ws[f'B{row}'] = plano_data.get('Impacto', 0)
    
    row += 1
    ws[f'A{row}'] = "Esforço"
    ws[f'B{row}'] = plano_data.get('Esforco', 0)
    
    row += 1
    ws[f'A{row}'] = "Score"
    ws[f'B{row}'] = plano_data.get('Score', 0)
    
    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 50
    
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# ============================================================================
# INTERFACE STREAMLIT
# ============================================================================

# Inicializar session state
if "plano_gerado" not in st.session_state:
    st.session_state.plano_gerado = None

if "historico_planos" not in st.session_state:
    st.session_state.historico_planos = []

# Header
st.markdown("""
<style>
    .header {
        background: linear-gradient(135deg, #1E3A8A 0%, #3B82F6 100%);
        padding: 2rem;
        border-radius: 10px;
        color: white;
        margin-bottom: 2rem;
    }
    .header h1 {
        margin: 0;
        font-size: 2.5rem;
    }
    .header p {
        margin: 0.5rem 0 0 0;
        font-size: 1.1rem;
        opacity: 0.9;
    }
</style>
<div class="header">
    <h1>📋 Diagnóstico 5W2H</h1>
    <p>Reunião de Start com Cliente - Gerador de Plano de Ação</p>
</div>
""", unsafe_allow_html=True)

# Abas principais
tab1, tab2, tab3 = st.tabs(["📝 Captura de Reunião", "📊 Plano 5W2H", "📈 Histórico"])

# ============================================================================
# ABA 1: CAPTURA DE REUNIÃO
# ============================================================================

with tab1:
    st.subheader("Preencha os dados da reunião de start")
    
    col1, col2 = st.columns(2)
    
    with col1:
        cliente = st.text_input("Nome do Cliente", placeholder="Ex: Empresa XYZ")
        responsavel = st.text_input("Responsável pela Execução", placeholder="Ex: João Silva")
    
    with col2:
        data_inicio = st.date_input("Data de Início", value=datetime.now())
        acao_selecionada = st.selectbox(
            "Selecione a Ação",
            options=[a["acao"] for a in BANCO_ACOES],
            help="Escolha a ação que será executada"
        )
    
    # Buscar ação selecionada
    acao_dict = obter_acao_por_nome(acao_selecionada)
    
    if acao_dict:
        st.markdown("---")
        st.subheader("Detalhes da Ação Selecionada")
        
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Categoria", acao_dict["categoria"])
        with col2:
            st.metric("Duração Padrão", f"{acao_dict['duracao_dias']} dias")
        with col3:
            st.metric("Impacto", f"{acao_dict['impacto_padrao']}/5")
        
        # Permitir customização
        st.markdown("---")
        st.subheader("Customizar Ação")
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            duracao_customizada = st.number_input(
                "Duração (dias)",
                value=acao_dict["duracao_dias"],
                min_value=1,
                max_value=60
            )
        
        with col2:
            impacto_customizado = st.slider(
                "Impacto (1-5)",
                min_value=1,
                max_value=5,
                value=acao_dict["impacto_padrao"]
            )
        
        with col3:
            esforco_customizado = st.slider(
                "Esforço (1-5)",
                min_value=1,
                max_value=5,
                value=acao_dict["esforco_padrao"]
            )
        
        # Notas da reunião
        notas = st.text_area(
            "Notas da Reunião",
            placeholder="Contexto, restrições, combinados, etc.",
            height=100
        )
        
        # Calcular score customizado
        score_customizado = calcular_score(impacto_customizado, esforco_customizado)
        
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Score Calculado", score_customizado)
        with col2:
            st.metric("Impacto × Peso", f"{impacto_customizado * CONFIG['peso_impacto']}")
        with col3:
            st.metric("Esforço × Peso", f"{esforco_customizado * CONFIG['peso_esforco']}")
        
        # Botão para gerar plano
        st.markdown("---")
        
        if st.button("🚀 Gerar Plano 5W2H", use_container_width=True, type="primary"):
            if not cliente or not responsavel:
                st.error("⚠️ Por favor, preencha o nome do cliente e responsável!")
            else:
                # Criar plano customizado
                plano_customizado = criar_dataframe_5w2h(
                    cliente, 
                    acao_dict, 
                    responsavel, 
                    notas,
                    data_inicio
                )
                
                # Aplicar customizações
                plano_customizado["Duracao_dias"] = duracao_customizada
                data_fim_customizada = data_inicio + timedelta(days=duracao_customizada)
                plano_customizado["Data_fim"] = data_fim_customizada.strftime("%d/%m/%Y")
                plano_customizado["Impacto"] = impacto_customizado
                plano_customizado["Esforco"] = esforco_customizado
                plano_customizado["Score"] = score_customizado
                
                # Salvar no session state
                st.session_state.plano_gerado = plano_customizado
                st.session_state.historico_planos.append(plano_customizado)
                
                st.success("✅ Plano 5W2H gerado com sucesso!")
                st.balloons()

# ============================================================================
# ABA 2: PLANO 5W2H
# ============================================================================

with tab2:
    if st.session_state.plano_gerado:
        plano = st.session_state.plano_gerado
        
        st.subheader(f"Plano 5W2H - {plano['Cliente']}")
        
        # Exibir em cards
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.metric("Cliente", plano["Cliente"])
        with col2:
            st.metric("Ação", plano["Acao"][:20] + "...")
        with col3:
            st.metric("Duração", f"{plano['Duracao_dias']} dias")
        with col4:
            st.metric("Score", plano["Score"])
        
        st.markdown("---")
        
        # 5W2H em abas
        w1, w2, w3, w4, w5, h1, h2 = st.tabs(["WHAT", "WHY", "WHERE", "WHEN", "WHO", "HOW", "HOW MUCH"])
        
        with w1:
            st.write("**O quê será feito?**")
            st.info(plano["What"])
        
        with w2:
            st.write("**Por quê fazer?**")
            st.info(plano["Why"])
        
        with w3:
            st.write("**Onde será feito?**")
            st.info(plano["Where"])
        
        with w4:
            st.write("**Quando será feito?**")
            st.write(f"**Início:** {plano['Data_inicio']}")
            st.write(f"**Fim:** {plano['Data_fim']}")
            st.write(f"**Duração:** {plano['Duracao_dias']} dias")
        
        with w5:
            st.write("**Quem vai fazer?**")
            st.info(plano["Responsavel"])
        
        with h1:
            st.write("**Como será feito?**")
            st.markdown(plano["How"])
        
        with h2:
            st.write("**Quanto custará?**")
            st.info(f"R$ {plano['HowMuch']:,.2f}")
        
        st.markdown("---")
        
        # Indicadores
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Impacto", f"{plano['Impacto']}/5")
        with col2:
            st.metric("Esforço", f"{plano['Esforco']}/5")
        with col3:
            st.metric("Score Final", plano['Score'])
        
        # Gráfico Gantt
        st.markdown("---")
        st.subheader("Timeline")
        fig_gantt = gerar_gantt_chart(plano)
        st.plotly_chart(fig_gantt, use_container_width=True)
        
        # Notas
        if plano.get("Notas"):
            st.markdown("---")
            st.subheader("Notas da Reunião")
            st.info(plano["Notas"])
        
        # Downloads
        st.markdown("---")
        st.subheader("Exportar Relatório")
        
        col1, col2 = st.columns(2)
        
        with col1:
            pdf_buffer = gerar_relatorio_pdf(plano["Cliente"], plano)
            st.download_button(
                label="📄 Baixar PDF",
                data=pdf_buffer,
                file_name=f"Plano_5W2H_{plano['Cliente']}_{datetime.now().strftime('%d%m%Y')}.pdf",
                mime="application/pdf"
            )
        
        with col2:
            excel_buffer = gerar_relatorio_excel(plano["Cliente"], plano)
            st.download_button(
                label="📊 Baixar Excel",
                data=excel_buffer,
                file_name=f"Plano_5W2H_{plano['Cliente']}_{datetime.now().strftime('%d%m%Y')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    
    else:
        st.info("👈 Preencha os dados na aba 'Captura de Reunião' para gerar o plano 5W2H")

# ============================================================================
# ABA 3: HISTÓRICO
# ============================================================================

with tab3:
    if st.session_state.historico_planos:
        st.subheader("Histórico de Planos Gerados")
        
        # Tabela com histórico
        df_historico = pd.DataFrame(st.session_state.historico_planos)
        df_exibicao = df_historico[["Cliente", "Acao", "Categoria", "Data_inicio", "Duracao_dias", "Score", "Status"]].copy()
        
        st.dataframe(df_exibicao, use_container_width=True, hide_index=True)
        
        # Estatísticas
        st.markdown("---")
        st.subheader("Estatísticas")
        
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.metric("Total de Planos", len(st.session_state.historico_planos))
        
        with col2:
            score_medio = df_historico["Score"].mean()
            st.metric("Score Médio", f"{score_medio:.1f}")
        
        with col3:
            impacto_medio = df_historico["Impacto"].mean()
            st.metric("Impacto Médio", f"{impacto_medio:.1f}/5")
        
        with col4:
            esforco_medio = df_historico["Esforco"].mean()
            st.metric("Esforço Médio", f"{esforco_medio:.1f}/5")
        
        # Gráficos
        st.markdown("---")
        
        col1, col2 = st.columns(2)
        
        with col1:
            fig_score = px.bar(
                df_historico,
                x="Acao",
                y="Score",
                title="Score por Ação",
                color="Score",
                color_continuous_scale="Blues"
            )
            st.plotly_chart(fig_score, use_container_width=True)
        
        with col2:
            fig_categoria = px.pie(
                df_historico,
                names="Categoria",
                title="Distribuição por Categoria"
            )
            st.plotly_chart(fig_categoria, use_container_width=True)
    
    else:
        st.info("Nenhum plano gerado ainda. Crie um novo plano na aba 'Captura de Reunião'")

# Footer
st.markdown("---")
st.markdown("""
<div style="text-align: center; color: #666; font-size: 0.9rem; margin-top: 2rem;">
    <p>Diagnóstico 5W2H v2.0 | Banco de Dados Otimizado</p>
    <p>Desenvolvido para reuniões de start com clientes</p>
    <p>Método: 5W2H (What, Why, Where, When, Who, How, How Much)</p>
</div>
""", unsafe_allow_html=True)
